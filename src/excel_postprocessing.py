#!/usr/bin/env python3

# Usage:
# excel_postprocessing.py cravat.tsv out.xlsx frequency_threshold [gene_panel] 

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import sys
import warnings

warnings.simplefilter("ignore")

def read_tsv(file):
    """
    Чтение TSV файла. Возвращает два DataFrame: один для данных, второй для шапки (метаинформации).
    """
    with open(file, 'r') as f:
        header_lines = []
        data_lines = []

        for line in f:
            if line.startswith('#'):
                header_lines.append(line.strip())
            else:
                data_lines.append(line.strip())
        
        # Преобразование шапки в DataFrame
        header_data = []
        for line in header_lines:
            # Убираем "Column description. " и "#", затем разделяем по "="
            line = line.replace("#", "").replace("Column description. ", "")
            if "=" in line:
                key, value = line.split("=", 1)
                header_data.append([key.strip(), value.strip()])
        
        header_df = pd.DataFrame(header_data, columns=["Тип данных", "Значение"])
        
        # Преобразование данных в DataFrame
        from io import StringIO
        data_str = '\n'.join(data_lines)
        data_df = pd.read_csv(StringIO(data_str), sep='\t')
        if 'uid' in data_df.columns:
            data_df = data_df.drop(columns=['uid'])

        # Переименование колонок
        rename_dict = {'chrom': 'Chrom', 'pos': 'Position', 'ref_base': 'Ref Base', 'alt_base': 'Alt Base', 'note_variant': 'Variant Note',
                       'coding': 'Coding', 'hugo': 'Gene', 'transcript': 'Transcript', 'so': 'Sequence Ontology', 'exonno': 'Exon Number',
                        'cchange': 'cDNA change', 'achange': 'Protein Change', 'all_mappings': 'All Mappings', 'gposend': 'End Position',
                        'numsample': 'Sample Count', 'samples': 'Samples', 'tags': 'Tags', 'hg19.chrom': 'hg19 Chrom', 'hg19.pos': 'hg19 Position',
                        'cardioboost.cardiomyopathy': 'CardioBoost Cardiomyopathy Score', 'cardioboost.cardiomyopathy1': 'CardioBoost Cardiomyopathy Class',
                        'cardioboost.arrhythmias': 'CardioBoost Arrhythmia Score', 'cardioboost.arrhythmias1': 'CardioBoost Arrhythmia Class',
                        'clinpred.score': 'ClinPred Score', 'clinpred.rankscore': 'ClinPred Rank Score', 'clinvar.sig': 'ClinVar Clinical Significance',
                        'clinvar.disease_refs': 'ClinVar Disease Ref Nums', 'clinvar.disease_refs_incl': 'ClinVar Disease Ref Nums Included Variant',
                        'clinvar.disease_names': 'ClinVar Disease Names', 'clinvar.clinvar_preferred_names': 'ClinVar Preferred Disease Names',
                        'clinvar.hgvs': 'ClinVar HGVS', 'clinvar.rev_stat': 'ClinVar Review Status', 'clinvar.id': 'ClinVar ClinVar ID',
                        'clinvar.sig_conf': 'ClinVar Significance Detail', 'clinvar.sig_conf_incl': 'ClinVar Significance Detail',
                        'clinvar.af_go_esp': 'ClinVar Allele Frequencies GO-ESP', 'clinvar.af_exac': 'ClinVar Allele Frequencies EXAC',
                        'clinvar.af_tgp': 'ClinVar Allele Frequencies TGP', 'clinvar.clinvar_allele_id': 'ClinVar Clinvar Allele Id',
                        'clinvar.variant_type': 'ClinVar Variant Type', 'clinvar.variant_type_sequence_ontology': 'ClinVar Variant Type SO',
                        'clinvar.variant_clinical_sources': 'ClinVar Variant Clinical Sources', 'clinvar.dbvar_id': 'ClinVar DBVAR ID',
                        'clinvar.clinvar_gene_info': 'ClinVar Gene names for the variant', 'clinvar.gene_info': 'ClinVar Molecular Consequence',
                        'clinvar.onc_disease_name': 'ClinVar Oncogenecity Disease Names', 'clinvar.onc_disease_name_incl': 'ClinVar Oncogenecity Disease Names Incl',
                        'clinvar.onc_disease_refs': 'ClinVar Oncogenecity Disease References',
                        'clinvar.onc_disease_refs_incl': 'ClinVar Oncogenecity Disease References Incl',
                        'clinvar.onc_classification': 'ClinVar Oncogenecity Classification',
                        'clinvar.onc_classification_type': 'ClinVar Oncogenecity Classification Type', 'clinvar.onc_rev_stat': 'ClinVar Oncogenecity Review Status',
                        'clinvar.onc_classification_conflicting': 'ClinVar Conflicting Oncogenecity Classification',
                        'clinvar.allele_origin': 'ClinVar Allele Origin', 'clinvar.dbsnp_id': 'ClinVar DBSNP ID',
                        'clinvar.somatic_disease_name': 'ClinVar Somatic Disease Name', 'clinvar.somatic_disease_name_incl': 'ClinVar Somatic Disease Name Incl',
                        'clinvar.somatic_refs': 'ClinVar Somatic References', 'clinvar.somatic_refs_incl': 'ClinVar Somatic References Incl',
                        'clinvar.somatic_rev_stat': 'ClinVar Somatic Review Status', 'clinvar.somatic_impact': 'ClinVar Somatic Impact',
                        'clinvar.somatic_impact_incl': 'ClinVar Somatic Impact Incl', 'clinvar.germline_or_somatic': 'ClinVar Germline or Somatic',
                        'clinvar_acmg.ps1_id': 'ClinVar ACMG PS1 ID', 'clinvar_acmg.pm5_id': 'ClinVar ACMG PM5 ID',
                        'extra_vcf_info.pos': 'Extra VCF INFO Annotations VCF Position', 'extra_vcf_info.ref': 'Extra VCF INFO Annotations VCF Ref Allele',
                        'extra_vcf_info.alt': 'Extra VCF INFO Annotations VCF Alt Allele', 'extra_vcf_info.NS': 'Extra VCF INFO Annotations NS',
                        'extra_vcf_info.DP': 'Extra VCF INFO Annotations DP', 'extra_vcf_info.DPB': 'Extra VCF INFO Annotations DPB',
                        'extra_vcf_info.AC': 'Extra VCF INFO Annotations AC', 'extra_vcf_info.AN': 'Extra VCF INFO Annotations AN',
                        'extra_vcf_info.AF': 'Extra VCF INFO Annotations AF', 'extra_vcf_info.RO': 'Extra VCF INFO Annotations RO',
                        'extra_vcf_info.AO': 'Extra VCF INFO Annotations AO', 'extra_vcf_info.PRO': 'Extra VCF INFO Annotations PRO',
                        'extra_vcf_info.PAO': 'Extra VCF INFO Annotations PAO', 'extra_vcf_info.QR': 'Extra VCF INFO Annotations QR',
                        'extra_vcf_info.QA': 'Extra VCF INFO Annotations QA', 'extra_vcf_info.PQR': 'Extra VCF INFO Annotations PQR',
                        'extra_vcf_info.PQA': 'Extra VCF INFO Annotations PQA', 'extra_vcf_info.SRF': 'Extra VCF INFO Annotations SRF',
                        'extra_vcf_info.SRR': 'Extra VCF INFO Annotations SRR', 'extra_vcf_info.SAF': 'Extra VCF INFO Annotations SAF',
                        'extra_vcf_info.SAR': 'Extra VCF INFO Annotations SAR', 'extra_vcf_info.SRP': 'Extra VCF INFO Annotations SRP',
                        'extra_vcf_info.SAP': 'Extra VCF INFO Annotations SAP', 'extra_vcf_info.AB': 'Extra VCF INFO Annotations AB',
                        'extra_vcf_info.ABP': 'Extra VCF INFO Annotations ABP', 'extra_vcf_info.RUN': 'Extra VCF INFO Annotations RUN',
                        'extra_vcf_info.RPP': 'Extra VCF INFO Annotations RPP', 'extra_vcf_info.RPPR': 'Extra VCF INFO Annotations RPPR',
                        'extra_vcf_info.RPL': 'Extra VCF INFO Annotations RPL', 'extra_vcf_info.RPR': 'Extra VCF INFO Annotations RPR',
                        'extra_vcf_info.EPP': 'Extra VCF INFO Annotations EPP', 'extra_vcf_info.EPPR': 'Extra VCF INFO Annotations EPPR',
                        'extra_vcf_info.DPRA': 'Extra VCF INFO Annotations DPRA', 'extra_vcf_info.ODDS': 'Extra VCF INFO Annotations ODDS',
                        'extra_vcf_info.GTI': 'Extra VCF INFO Annotations GTI', 'extra_vcf_info.TYPE': 'Extra VCF INFO Annotations TYPE',
                        'extra_vcf_info.CIGAR': 'Extra VCF INFO Annotations CIGAR', 'extra_vcf_info.NUMALT': 'Extra VCF INFO Annotations NUMALT',
                        'extra_vcf_info.MEANALT': 'Extra VCF INFO Annotations MEANALT', 'extra_vcf_info.LEN': 'Extra VCF INFO Annotations LEN',
                        'extra_vcf_info.MQM': 'Extra VCF INFO Annotations MQM', 'extra_vcf_info.MQMR': 'Extra VCF INFO Annotations MQMR',
                        'extra_vcf_info.PAIRED': 'Extra VCF INFO Annotations PAIRED', 'extra_vcf_info.PAIREDR': 'Extra VCF INFO Annotations PAIREDR',
                        'extra_vcf_info.MIN_DP': 'Extra VCF INFO Annotations MIN_DP', 'extra_vcf_info.END': 'Extra VCF INFO Annotations END',
                        'hpo.id': 'Human Phenotype Ontology HPO ID', 'hpo.term': 'Human Phenotype Ontology HPO Term',
                        'hpo.all': 'Human Phenotype Ontology All Annotations', 'ncbigene.ncbi_desc': 'NCBI Gene Description', 'ncbigene.entrez': 'NCBI Gene Entrez ID',
                        'omim.omim_id': 'OMIM Entry ID', 'original_input.chrom': 'Original Input Chrom', 'original_input.pos': 'Original Input Pos',
                        'original_input.ref_base': 'Original Input Reference allele', 'original_input.alt_base': 'Original Input Alternate allele',
                        'spliceai.ds_ag': 'SpliceAI Acceptor Gain Score', 'spliceai.ds_al': 'SpliceAI Acceptor Loss Score', 'spliceai.ds_dg': 'SpliceAI Donor Gain Score',
                        'spliceai.ds_dl': 'SpliceAI Donor Loss Score', 'spliceai.dp_ag': 'SpliceAI Acceptor Gain Position',
                        'spliceai.dp_al': 'SpliceAI Acceptor Loss Position', 'spliceai.dp_dg': 'SpliceAI Donor Gain Position',
                        'spliceai.dp_dl': 'SpliceAI Donor Loss Posiiton', 'vcfinfo.phred': 'VCF Info Phred', 'vcfinfo.filter': 'VCF Info VCF filter',
                        'vcfinfo.zygosity': 'VCF Info Zygosity', 'vcfinfo.alt_reads': 'VCF Info Alternate reads', 'vcfinfo.tot_reads': 'VCF Info Total reads',
                        'vcfinfo.af': 'VCF Info Variant AF', 'vcfinfo.hap_block': 'VCF Info Haplotype block ID', 'vcfinfo.hap_strand': 'VCF Info Haplotype strand ID',
                        'gnomad.af': 'gnomAD Global AF', 'gnomad.af_afr': 'gnomAD African AF', 'gnomad.af_amr': 'gnomAD American AF',
                        'gnomad.af_asj': 'gnomAD Ashkenazi Jewish AF', 'gnomad.af_eas': 'gnomAD East Asian AF', 'gnomad.af_fin': 'gnomAD Finnish AF',
                        'gnomad.af_nfe': 'gnomAD Non-Fin Eur AF', 'gnomad.af_oth': 'gnomAD Other AF', 'gnomad.af_sas': 'gnomAD South Asian AF',
                        'gnomad4.af': 'gnomAD4 Global AF', 'gnomad4.af_afr': 'gnomAD4 African AF', 'gnomad4.af_ami': 'gnomAD4 Amish AF',
                        'gnomad4.af_amr': 'gnomAD4 American AF', 'gnomad4.af_asj': 'gnomAD4 Ashkenazi Jewish AF', 'gnomad4.af_eas': 'gnomAD4 East Asian AF',
                        'gnomad4.af_fin': 'gnomAD4 Finnish AF', 'gnomad4.af_mid': 'gnomAD4 Mid-East AF', 'gnomad4.af_nfe': 'gnomAD4 Non-Fin Eur AF',
                        'gnomad4.af_sas': 'gnomAD4 South Asian AF'}  
        data_df.rename(columns=rename_dict, inplace=True)

    return data_df, header_df


def create_excels(data_df:pd.DataFrame, header_df:pd.DataFrame, output_file:str, var_threshold:float, gene_panel:list=[]):
    """
    Создает 2 Excel-файла (один со всеми частотами, другой (клинический) - с частотами не более указанной).\n
    Файлы с двумя листами: один для данных, другой для шапки.
    """

    dfs = reform_data(data_df, var_threshold, output_file, gene_panel)

    for filepath, df in dfs.items():
        wb = add_header_sheet(design_data_df(df), header_df)
        # Сохранение файла
        try:
            wb.save(filepath)
        except ValueError:
            print(f"Can't save {filepath}: too big!")
            continue


def reform_data(data_df:pd.DataFrame, var_threshold:float, output_file:str, gene_panel:list) -> dict:
    add_varsome_column(data_df)
    # Перемещение большинства колонок 'Extra VCF Info, Original Input' в конец;  важных из Extra VCF Info - в начало
    important_cols = ['SRF', 'SRR', 'SAF', 'SAR']
    alt_base_idx = data_df.columns.get_loc('Alt Base') + 1
    df_cols = data_df.columns.values
    for col in df_cols:
        if 'Extra VCF INFO' in col:
            if any(important_col in col for important_col in important_cols):
                data_df.insert(alt_base_idx + 1, col, data_df.pop(col))
            else:
                extra_vcf_info_col = data_df.pop(col)
                data_df[col] = extra_vcf_info_col
        if 'Original Input' in col:
            original_input_col = data_df.pop(col)
            data_df[col] = original_input_col
    # перемещение налево Gnomad Global AF
    for col in ['gnomAD Global AF', 'gnomAD4 Global AF']:
        if col in data_df.columns.values:
            data_df.insert(alt_base_idx + 1, col, data_df.pop(col))
    # AF ставим всё же левее всех остальных
    data_df.insert(alt_base_idx + 1, 'Extra VCF INFO Annotations AF', data_df.pop('Extra VCF INFO Annotations AF'))
    clinical_data_df = data_df[data_df['gnomAD Global AF'] <= var_threshold].reset_index(drop=True)
    if gene_panel:
        gene_panel_df = clinical_data_df[clinical_data_df['Gene'].isin(gene_panel)].reset_index(drop=True)
        return {output_file:data_df,
                output_file.replace('.xlsx', '_clinical.xlsx'):clinical_data_df,
                output_file.replace('.xlsx', '_panel_clinical.xlsx'):gene_panel_df}
    return {output_file:data_df, output_file.replace('.xlsx', '_clinical.xlsx'):clinical_data_df}

def add_varsome_column(data_df):
    """
    Добавляет новый столбец 'Varsome' в DataFrame, который содержит ссылки на varsome 
    для каждого варианта, основанные на значениях в столбцах координат.
    """
    # Создаём новый столбец для VarSome
    data_df['Varsome'] = ''

    # Итерация по строкам DataFrame
    for index, row in data_df.iterrows():
        var_ids = []
        # Объединяем значения из столбцов координат, пропуская "-"
        for col in range(4):
            if row[col] != "-":
                var_ids.append(str(row[col]))
            else:
                if col == 2:
                    var_ids.append('-')
        
        # Создаём строку из идентификаторов, разделённых дефисами
        if var_ids:
            var_string = "-".join(var_ids).replace('---', '--')
            # Формируем ссылку на varsome
            data_df.at[index, 'Varsome'] = f"https://varsome.com/variant/hg38/{var_string}?annotation-mode=germline"


def design_data_df(df:pd.DataFrame) -> Workbook:
    wb = Workbook()
    
    # Лист для данных
    ws_data = wb.active
    ws_data.title = "Данные"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    # Применение цветовой кодировки к Gnomad Global AF
    for col in ['gnomAD Global AF', 'gnomAD4 Global AF']:
        if col in df.columns.values:
            global_af_colors = df[col].apply(colorize_global_af)
            # Добавляем форматы в Excel после создания листа
            for idx, color in enumerate(global_af_colors, start=2):  # Индексация с 2, чтобы пропустить заголовок
                cell = ws_data.cell(row=idx, column=df.columns.get_loc(col) + 1)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    format_dataframe(ws_data, df)
    # Добавление проверки данных (пример для первой колонки)
    #add_data_validation(ws_data, df, 'A')
    return wb

def colorize_global_af(val:float) -> str:
    """Возвращает цвет в зависимости от значения 'Global AF'."""
    if pd.isnull(val) or val < 0.01:
        return "FF0000"  # Красный
    elif 0.01 <= val <= 0.05:
        return "FFFF00"  # Жёлтый
    elif val > 0.05:
        return "00FF00"  # Зелёный
    return "FFFFFF"  # Белый для пустых значений

def format_dataframe(ws, df, max_width=15):
    """
    Форматирует DataFrame в Excel: добавляет жирный шрифт к заголовкам, 
    устанавливает оптимальную ширину колонок, добавляет автофильтр и
    прибавляет 3 символа к ширине колонок для учёта кнопки автофильтра.
    """
    for idx, col in enumerate(df.columns):
        # Жирный шрифт для заголовков, выравнивание по центру
        cell = ws.cell(row=1, column=idx+1)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

        # Определяем оптимальную ширину для каждой колонки
        series = df[col].astype(str)
        try:
            max_len = max(series.map(len).max(), len(col)) + 6
        except ValueError:
            max_len = 10
        # Если длина текста больше максимальной ширины, устанавливаем максимальную ширину
        column_width = min(max_len + 2, max_width)  # Добавляем 2 символа для лучшей читабельности
        ws.column_dimensions[get_column_letter(idx+1)].width = column_width

    # Добавляем автофильтр для всех колонок
    ws.auto_filter.ref = ws.dimensions

def add_header_sheet(wb:Workbook, header_df:pd.DataFrame) -> Workbook:
    # Лист для шапки
    ws_header = wb.create_sheet(title="Описание")
    
    for r in dataframe_to_rows(header_df, index=False, header=True):
        ws_header.append(r)
    
    format_dataframe(ws_header, header_df)

    return wb


def main():
    """
    Основная функция. Читает входной TSV файл, преобразует его и сохраняет в Excel.
    """
    if len(sys.argv) not in [4,5]:
        print("Использование: excel_postprocessing.py cravat.tsv out.xlsx frequency_threshold [gene_panel]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]
    var_threshold = float(sys.argv[3])
    gene_panel = []
    if len(sys.argv) == 5:
        gene_panel = pd.read_excel(sys.argv[4])['Gene'].to_list()
        
    # Чтение TSV файла
    data_df, header_df = read_tsv(input_file)

    # Создание Excel-файлов
    create_excels(data_df=data_df, header_df=header_df, output_file=output_file, var_threshold=var_threshold, gene_panel=gene_panel)


if __name__ == "__main__":
    main()